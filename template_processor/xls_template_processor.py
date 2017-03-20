# coding: utf-8

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple
from openpyxl.utils import get_column_letter
import pandas as pd
from PySide.QtCore import *

from template_processor.array_template_processor import ArrayTemplateProcessor


class XlsTemplateProcessor(ArrayTemplateProcessor):

    def __init__(self, xls_template_path, fuchia_database):
        self._xls_template_path = xls_template_path
        self._workbook = load_workbook(self.xls_template_path)
        super(XlsTemplateProcessor, self).__init__(
            pd.DataFrame(self._workbook.active.values).as_matrix(),
            fuchia_database
        )

    @property
    def xls_template_path(self):
        return self._xls_template_path

    @xls_template_path.setter
    def xls_template_path(self, value):
        self._xls_template_path = value
        self._workbook = load_workbook(self.xls_template_path)
        self.array = pd.DataFrame(self._workbook.active.values).as_matrix()

    def get_merged_cell_ranges(self):
        """
        :return: A list containing the ranges of merged cells. A range is
        composed with two couples, describing the up left cell and the down
        right cell. e.g. ((1, 1), (1, 2)).
        """
        merged_cell_ranges = self._workbook.active.merged_cell_ranges
        split = [i.split(':') for i in merged_cell_ranges]
        coordinates = [tuple([coordinate_to_tuple(i) for i in t])
                       for t in split]
        fixed = []
        # Fix indices
        for r in coordinates:
            a, b = r[0]
            c, d = r[1]
            fixed.append(((a - 1, b - 1), (c - 1, d - 1)))
        return fixed

    def get_cell_style(self, i, j):
        """
        :return: If a style is available for a given cell, return a dict,
        with style information about font, alignment fill and stroke.
        """
        cell = self._workbook.active.cell(row=i + 1, column=j + 1)
        style = {}
        if cell.fill:
            rgba_hex = cell.fill.start_color.rgb
            a = int(rgba_hex[:2], 16)
            r = int(rgba_hex[2:4], 16)
            g = int(rgba_hex[4:6], 16)
            b = int(rgba_hex[6:8], 16)
            style['fill'] = {
                'color': (r, g, b, a),
            }
        if cell.alignment:
            h_align = cell.alignment.horizontal
            v_align = cell.alignment.vertical
            style['alignment'] = get_align_flag(h_align, v_align)
        return style

    def get_column_width(self, j):
        col_letter = get_column_letter(j + 1)
        w = self._workbook.active.column_dimensions[col_letter].width
        if w is None:
            return w
        return w * 10

    def get_row_height(self, i):
        h = self._workbook.active.row_dimensions[i + 1].height
        if h is None:
            return h
        return h * 2

    def export_to_excel(self, destination_path):
        wb = load_workbook(self.xls_template_path)
        ws = wb.active
        for key, value in self.last_template_values.items():
            i, j = key[0], key[1]
            cell = ws.cell(row=i + 1, column=j + 1)
            cell.value = value
        wb.save(destination_path)


def get_align_flag(h_align, v_align):
    h_map = {
        'left': Qt.AlignLeft,
        'center': Qt.AlignHCenter,
        'right': Qt.AlignRight,
        'justify': Qt.AlignJustify,
    }
    v_map = {
        'top': Qt.AlignTop,
        'bottom': Qt.AlignBottom,
        'center': Qt.AlignVCenter,
    }
    h_align_flag = Qt.AlignLeft
    if h_align in h_map:
        h_align_flag = h_map[h_align]
    v_align_flag = Qt.AlignTop
    if v_align in v_map:
        v_align_flag = v_map[v_align]
    return h_align_flag | v_align_flag
