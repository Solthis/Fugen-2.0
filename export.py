# -*- coding: utf-8 -*

'''
Export module.

@author: Dimitri Justeau <dimitri.justeau@gmail.com>
'''

from openpyxl.reader.excel import load_workbook
from openpyxl.utils.cell import get_column_letter

import reports.medical.indicators as medkeys
import constants


# Medical xlsx report constants
E_VALUES_RANGE = [(9, 16), (3, 7)]
A_VALUES_RANGE = [(17, 21), (3, 7)]
S_VALUES_RANGE = [(22, 26), (3, 7)]
B_VALUES_RANGE = [(27, 29), (3, 7)]
PEC_ENTITY_CELL = 'B3'
PEC_PERIOD_CELL = 'B4'


def exportMedicalReportToExcel(center, month, year, indicators, dest_path):
    wb = load_workbook(filename=constants.MEDICAL_REPORT_TEMPLATE)
    ws = wb.active
    ws.cell(PEC_ENTITY_CELL).value = center
    p = '{}/{}'.format(str(month).zfill(2), year)
    ws.cell(PEC_PERIOD_CELL).value = p
    INDICS = [medkeys.E, medkeys.A, medkeys.S, medkeys.B]
    for k, RANGE in enumerate([E_VALUES_RANGE, A_VALUES_RANGE, S_VALUES_RANGE, B_VALUES_RANGE]):
        i = 0
        for row in range(RANGE[0][0], RANGE[0][1]):
            indic = INDICS[k][i]
            c = 0
            for col_idx in range(RANGE[1][0], RANGE[1][1]):
                cat = medkeys.INDICATOR_CATEGORY_KEYS[c]
                col = get_column_letter(col_idx)
                value = indicators[cat][indic]
                ws.cell('{}{}'.format(col, row)).value = value
                c += 1
            i += 1
    # Save the excel sheet
    wb.save(filename=dest_path)
